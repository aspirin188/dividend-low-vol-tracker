"""
路由 — 红利低波跟踪系统

v8.0 新增：
  GET  /api/config/strategies   → 获取预设策略列表
  POST /api/config/strategies  → 应用预设策略
  GET  /api/config/history     → 获取参数历史记录

v6.20 新增配置管理：
  GET  /config           → 配置管理页面
  GET  /api/config       → 获取所有配置
  PUT  /api/config/batch → 批量更新配置
  POST /api/config/reset → 恢复默认值

原有 API：
  GET  /          → 渲染首页
  POST /api/run   → 一键运行（同步，超时10分钟）
  GET  /api/stocks→ 获取标的列表（支持搜索和行业筛选）
  GET  /api/export→ 导出 Excel
"""

import os
import sqlite3
import pandas as pd
from io import BytesIO
from flask import (
    Blueprint, render_template, request, jsonify, send_file,
    current_app,
)
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, numbers
from openpyxl.utils import get_column_letter

from server.services.fetcher import merge_all_data
from server.services.scorer import filter_stocks, calculate_scores, prepare_results
from server.services.health_tracker import get_health_tracker  # v6.19新增
from server.services.config_service import ConfigService  # v6.20新增

bp = Blueprint('main', __name__)


# ============================================================
# 数据库辅助
# ============================================================

def get_db():
    """获取数据库连接。"""
    db_path = os.path.join(current_app.instance_path, 'tracker.db')
    os.makedirs(os.path.dirname(db_path), exist_ok=True)
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    """初始化数据库表。"""
    conn = get_db()
    conn.execute('''
        CREATE TABLE IF NOT EXISTS stock_data (
            code             TEXT PRIMARY KEY,
            name             TEXT,
            industry         TEXT,
            market           TEXT,
            dividend_yield   REAL,
            annual_vol       REAL,
            composite_score  REAL,
            rank             INTEGER,
            market_cap       REAL,
            payout_ratio     REAL,
            eps              REAL,
            price            REAL,
            pe               REAL,
            pb               REAL,
            pinyin_abbr      TEXT,
            data_date        TEXT,
            updated_at       TEXT
        )
    ''')

    # 兼容旧数据库：新增列
    for col, col_type in [
        ('price', 'REAL'), ('pe', 'REAL'), ('pb', 'REAL'), 
        ('market', 'TEXT'), ('pinyin_abbr', 'TEXT'), 
        ('dividend_years', 'INTEGER'), ('roe', 'REAL'), ('debt_ratio', 'REAL'),
        ('price_percentile', 'REAL'),  # v6.19新增
        ('payout_3y_avg', 'REAL'),  # v7.0新增
        # v7.2新增：质量因子增强
        ('profit_growth_3y', 'REAL'),
        ('cashflow_profit_ratio', 'REAL'),
        ('top1_shareholder_ratio', 'REAL'),
        ('strike_zone_score', 'REAL'),
        ('strike_zone_rating', 'TEXT'),
        ('strike_zone', 'TEXT'),
        # v7.2.1新增：均线位置
        ('ma250', 'REAL'),
        ('ma20', 'REAL'),  # v7.3新增
        ('ma60', 'REAL'),  # v7.3新增
        ('current_price', 'REAL'),  # v7.3新增
        ('price_vs_ma_pct', 'REAL'),
        ('ma_slope', 'REAL'),
        ('trend', 'TEXT'),  # v7.3新增：趋势方向
        ('trend_strength', 'TEXT'),  # v7.3新增：趋势强度
        ('signal', 'TEXT'),
        ('signal_level', 'INTEGER'),
        ('signal_type', 'TEXT'),  # v7.3新增：信号类型 buy/sell/hold
        ('action', 'TEXT'),  # v7.3新增：操作建议
        ('ma_score', 'REAL'),
        # v8.4新增：成长因子
        ('growth_factor', 'REAL'),
        ('peg', 'REAL'),
        ('roe_trend', 'REAL'),
    ]:
        try:
            conn.execute(f'ALTER TABLE stock_data ADD COLUMN {col} {col_type}')
        except sqlite3.OperationalError:
            pass  # 列已存在

    conn.commit()
    conn.close()


# ============================================================
# GET / → 首页
# ============================================================

@bp.route('/')
def index():
    return render_template('index.html')


# ============================================================
# POST /api/run → 一键运行
# ============================================================

@bp.route('/api/run', methods=['POST'])
def run():
    """
    一键运行：获取数据 → 筛选 → 评分 → 保存。
    同步执行，前端超时设为 600s。
    
    v7.2更新：
    - 新增质量因子数据获取（净利润增速、现金流、股东持股）
    - 新增击球区评分计算
    """
    try:
        # 1. 合并所有数据
        merged = merge_all_data()

        if merged.empty:
            return jsonify({'success': False, 'error': '未获取到任何数据，请检查网络'})

        # v7.2新增：获取质量因子增强数据
        print("步骤9: 获取分红年数数据（v8.2修复）...")
        # v8.4.1修复：profit_growth_3y已通过fetch_profit_growth_data获取，不需要重复查询
        from server.services.fetcher import (
            # get_profit_history_batch,  # v8.4.1删除：profit_growth_3y已通过fetch_profit_growth_data获取
            get_operating_cashflow_batch,
            get_top_shareholder_ratio_batch,
            # calculate_profit_growth_3y,  # v8.4.1删除：已通过fetch_profit_growth_data获取
            calculate_cashflow_profit_ratio,
            calc_ma_position_batch,  # v7.2.1新增
            # is_profit_growing_strict,  # v8.4.1删除：不再使用
            get_dividend_years_batch,  # v8.2修复
            fetch_profit_growth_data,  # v8.4新增
        )
        
        candidate_codes = merged['code'].tolist()
        
        # v8.2修复：获取真实分红年数
        dividend_years_map = get_dividend_years_batch(candidate_codes, years=4)
        merged['dividend_years'] = merged['code'].map(lambda x: dividend_years_map.get(x, 0))
        
        # v8.4新增：获取成长因子数据（净利润增长率、ROE趋势）
        print("步骤9.1: 获取成长因子数据（v8.4）...")
        growth_data = fetch_profit_growth_data(candidate_codes, years=3)
        merged['profit_growth_3y'] = merged['code'].map(lambda x: growth_data.get(x, {}).get('profit_growth_3y'))
        merged['roe_trend'] = merged['code'].map(lambda x: growth_data.get(x, {}).get('roe_trend'))
        merged['peg'] = None  # PEG在scorer中计算
        merged['growth_factor'] = None  # growth_factor在scorer中计算

        # v8.4.1新增：利润增长筛选（过滤负增长股票）
        min_growth = config.get_float('MIN_PROFIT_GROWTH', default=0)
        enable_filter = config.get_bool('ENABLE_PROFIT_GROWTH_FILTER')
        if enable_filter:
            before_count = len(merged)
            # 无数据放行，负增长过滤
            mask = merged['profit_growth_3y'].isna() | (merged['profit_growth_3y'] >= min_growth)
            merged = merged[mask].copy()
            filtered_count = before_count - len(merged)
            if filtered_count > 0:
                print(f"  ✓ 利润增长筛选过滤{filtered_count}只（负增长）")
        
        # 1.2 获取经营现金流数据
        operating_cashflow = get_operating_cashflow_batch(candidate_codes)
        
        # 1.3 获取第一大股东持股比例
        top_shareholder_ratio = get_top_shareholder_ratio_batch(candidate_codes)
        
        # 1.4 计算均线位置（v7.2.1新增）
        print("步骤12: 计算均线位置和买点信号...")
        ma_data = calc_ma_position_batch(candidate_codes)
        
        # v8.4.1修复：保留fetch_profit_growth_data获取的成长因子数据
        # 删除merged['profit_growth_3y'] = None，避免覆盖v8.4新增的数据
        # 注意：profit_growth_3y和roe_trend已在第174-176行获取，这里不能覆盖
        merged['cashflow_profit_ratio'] = None
        merged['top1_shareholder_ratio'] = None
        merged['ma250'] = None  # v7.2.1新增
        merged['ma20'] = None  # v7.3新增
        merged['ma60'] = None  # v7.3新增
        merged['current_price'] = None  # v7.3新增
        merged['price_vs_ma_pct'] = None  # v7.2.1新增
        merged['ma_slope'] = None  # v7.2.1新增
        merged['trend'] = None  # v7.3新增
        merged['trend_strength'] = None  # v7.3新增
        merged['signal'] = None  # v7.2.1新增
        merged['signal_level'] = None  # v7.2.1新增
        merged['signal_type'] = None  # v7.3新增
        merged['action'] = None  # v7.3新增
        
        for code in candidate_codes:
            # v8.4.1修复：profit_growth_3y已在第174-176行通过fetch_profit_growth_data获取
            # 不需要在此处重复计算，避免数据不一致
            
            # 计算现金流质量
            if code in operating_cashflow:
                # 从merged中获取净利润
                net_profit = merged.loc[merged['code'] == code, 'basic_eps'].values[0]
                if pd.notna(net_profit) and net_profit > 0:
                    ratio = calculate_cashflow_profit_ratio(operating_cashflow[code], net_profit * 1e8)  # EPS转净利润
                    if ratio is not None:
                        merged.loc[merged['code'] == code, 'cashflow_profit_ratio'] = ratio
            
            # 第一大股东持股比例
            if code in top_shareholder_ratio:
                merged.loc[merged['code'] == code, 'top1_shareholder_ratio'] = top_shareholder_ratio[code]
            
            # 均线位置数据（v7.2.1新增，v7.3升级）
            if code in ma_data:
                merged.loc[merged['code'] == code, 'ma250'] = ma_data[code]['ma250']
                merged.loc[merged['code'] == code, 'ma20'] = ma_data[code]['ma20']  # v7.3新增
                merged.loc[merged['code'] == code, 'ma60'] = ma_data[code]['ma60']  # v7.3新增
                merged.loc[merged['code'] == code, 'current_price'] = ma_data[code]['current_price']  # v7.3新增
                merged.loc[merged['code'] == code, 'price_vs_ma_pct'] = ma_data[code]['price_vs_ma_pct']
                merged.loc[merged['code'] == code, 'ma_slope'] = ma_data[code]['ma_slope']
                merged.loc[merged['code'] == code, 'trend'] = ma_data[code]['trend']  # v7.3新增
                merged.loc[merged['code'] == code, 'trend_strength'] = ma_data[code]['trend_strength']  # v7.3新增
                merged.loc[merged['code'] == code, 'signal'] = ma_data[code]['signal']
                merged.loc[merged['code'] == code, 'signal_level'] = ma_data[code]['signal_level']
                merged.loc[merged['code'] == code, 'signal_type'] = ma_data[code]['signal_type']  # v7.3新增
                merged.loc[merged['code'] == code, 'action'] = ma_data[code]['action']  # v7.3新增
        
        print(f"  ✓ 成功获取质量因子数据")

        # 2. 硬性筛选
        filtered = filter_stocks(merged)
        print(f"筛选后: {len(filtered)} 只")

        if filtered.empty:
            return jsonify({'success': False, 'error': '没有符合条件的股票，请稍后再试'})

        # 3. 评分
        scored = calculate_scores(filtered)
        
        # v7.2新增：计算击球区评分
        from server.services.scorer import calculate_strike_zone_score
        scored = calculate_strike_zone_score(scored)

        # 4. 整理并保存
        result = prepare_results(scored)
        conn = get_db()

        # 清空旧数据
        conn.execute('DELETE FROM stock_data')

        # 插入新数据
        result.to_sql('stock_data', conn, if_exists='append', index=False)
        conn.commit()
        conn.close()

        data_date = result['data_date'].iloc[0] if not result.empty else ''
        return jsonify({
            'success': True,
            'count': len(result),
            'data_date': data_date,
        })

    except Exception as e:
        print(f"运行失败: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)})


# ============================================================
# GET /api/stocks → 标的列表
# ============================================================

@bp.route('/api/stocks')
def stocks():
    """
    获取标的列表，支持搜索和多维筛选。

    Query 参数：
    - q: 搜索关键词（名称或代码）
    - industry: 行业筛选（大类行业）
    - market_cap_range: 市值区间（large≥1000亿 / medium 500-1000亿 / small 100-500亿）
    - div_yield_min: 股息率下限（4 / 6 / 8）
    - market: 市场类型（sh=沪市 / sz=深市 / gem=创业板 / star=科创板）
    
    v6.19新增：
    - roe_min: ROE下限（8 / 10 / 12 / 15）
    - debt_ratio_max: 负债率上限（60 / 70 / 80 / 85）
    - payout_min: 支付率下限（30 / 50 / 70）
    - price_percentile_max: 股价百分位上限（20 / 40 / 60 / 80）
    """
    q = request.args.get('q', '').strip()
    industry = request.args.get('industry', '').strip()
    market_cap_range = request.args.get('market_cap_range', '').strip()
    div_yield_min = request.args.get('div_yield_min', '').strip()
    market = request.args.get('market', '').strip()
    
    # v6.19: 高级筛选参数
    roe_min = request.args.get('roe_min', '').strip()
    debt_ratio_max = request.args.get('debt_ratio_max', '').strip()
    payout_min = request.args.get('payout_min', '').strip()
    price_percentile_max = request.args.get('price_percentile_max', '').strip()

    conn = get_db()
    conn.row_factory = sqlite3.Row
    query = 'SELECT * FROM stock_data ORDER BY rank'
    params = []

    conditions = []
    if q:
        # v6.9: 支持简拼搜索，同时匹配名称、代码、拼音首字母
        conditions.append('(name LIKE ? OR code LIKE ? OR pinyin_abbr LIKE ?)')
        params.extend([f'%{q}%', f'%{q}%', f'%{q}%'])
    if industry:
        conditions.append('industry = ?')
        params.append(industry)
    if market_cap_range:
        cap_ranges = {
            'large': ('market_cap >= 1000', []),
            'medium': ('market_cap >= 500 AND market_cap < 1000', []),
            'small': ('market_cap >= 100 AND market_cap < 500', []),
        }
        if market_cap_range in cap_ranges:
            conditions.append(cap_ranges[market_cap_range][0])
            params.extend(cap_ranges[market_cap_range][1])
    if div_yield_min:
        try:
            min_yield = float(div_yield_min)
            conditions.append('dividend_yield >= ?')
            params.append(min_yield)
        except ValueError:
            pass
    if market:
        market_map = {
            'sh': '沪市',
            'sz': '深市',
            'gem': '创业板',
            'star': '科创板',
        }
        if market in market_map:
            conditions.append('market = ?')
            params.append(market_map[market])
    
    # v6.19: 高级筛选
    if roe_min:
        try:
            min_roe = float(roe_min)
            conditions.append('roe >= ?')
            params.append(min_roe)
        except ValueError:
            pass
    
    if debt_ratio_max:
        try:
            max_debt = float(debt_ratio_max)
            conditions.append('debt_ratio <= ?')
            params.append(max_debt)
        except ValueError:
            pass
    
    if payout_min:
        try:
            min_payout = float(payout_min)
            conditions.append('payout_ratio >= ?')
            params.append(min_payout)
        except ValueError:
            pass
    
    if price_percentile_max:
        try:
            max_percentile = float(price_percentile_max)
            conditions.append('price_percentile <= ?')
            params.append(max_percentile)
        except ValueError:
            pass

    if conditions:
        query = 'SELECT * FROM stock_data WHERE ' + ' AND '.join(conditions) + ' ORDER BY rank'

    rows = conn.execute(query, params).fetchall()

    # 获取行业列表（用于筛选下拉）
    industries = [r['industry'] for r in conn.execute(
        'SELECT DISTINCT industry FROM stock_data WHERE industry IS NOT NULL AND industry != "" ORDER BY industry'
    ).fetchall()]

    # 获取元数据
    meta = conn.execute('SELECT data_date, updated_at FROM stock_data LIMIT 1').fetchone()

    conn.close()

    stock_list = [dict(r) for r in rows]
    
    # v6.19: 获取健康度数据
    health_tracker = get_health_tracker()
    health_data = {
        'sina_reliability': health_tracker.get_reliability('sina_price'),
        'tencent_reliability': health_tracker.get_reliability('tencent_price'),
    }

    return jsonify({
        'stocks': stock_list,
        'industries': industries,
        'total': len(stock_list),
        'data_date': dict(meta)['data_date'] if meta else None,
        'updated_at': dict(meta)['updated_at'] if meta else None,
        'health': health_data,  # v6.19新增
    })


# ============================================================
# GET /api/export → 导出 Excel
# ============================================================

@bp.route('/api/export')
def export():
    """
    导出 Excel。

    文件名：红利低波_{data_date}_{count}只.xlsx
    列（17列）：排名 | 代码 | 名称 | 行业 | 市场 | 股价 | PE | PB | 股息率(%) | 总市值(亿) | 综合评分 | 波动率(%) | 股利支付率(%) | EPS | 分红年数 | ROE(%) | 负债率(%)
    """
    conn = get_db()
    conn.row_factory = sqlite3.Row
    rows = conn.execute('SELECT * FROM stock_data ORDER BY rank').fetchall()
    conn.close()

    if not rows:
        return jsonify({'error': '没有数据可导出，请先运行'}), 404

    stocks = [dict(r) for r in rows]
    data_date = stocks[0].get('data_date', 'unknown')
    count = len(stocks)

    # 创建 Excel
    wb = Workbook()
    ws = wb.active
    ws.title = '红利低波'

    # 表头（17列）- v6.11新增ROE、负债率
    headers = ['排名', '代码', '名称', '行业', '市场', '股价', 'PE', 'PB', '股息率(%)', '总市值(亿)', '综合评分', '波动率(%)', '股利支付率(%)', 'EPS', '分红年数', 'ROE(%)', '负债率(%)']
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal='center')

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.alignment = header_alignment

    # 数据行
    for row_idx, s in enumerate(stocks, 2):
        ws.cell(row=row_idx, column=1, value=s.get('rank')).alignment = Alignment(horizontal='center')
        ws.cell(row=row_idx, column=2, value=s.get('code')).alignment = Alignment(horizontal='center')
        ws.cell(row=row_idx, column=3, value=s.get('name'))
        ws.cell(row=row_idx, column=4, value=s.get('industry') or '')
        ws.cell(row=row_idx, column=5, value=s.get('market') or '')
        ws.cell(row=row_idx, column=6, value=s.get('price')).number_format = '0.00'
        ws.cell(row=row_idx, column=7, value=s.get('pe')).number_format = '0.00'
        ws.cell(row=row_idx, column=8, value=s.get('pb')).number_format = '0.00'
        ws.cell(row=row_idx, column=9, value=s.get('dividend_yield')).number_format = '0.00'
        ws.cell(row=row_idx, column=10, value=s.get('market_cap')).number_format = '0.00'
        ws.cell(row=row_idx, column=11, value=s.get('composite_score')).number_format = '0.00'
        ws.cell(row=row_idx, column=12, value=s.get('annual_vol')).number_format = '0.00'
        ws.cell(row=row_idx, column=13, value=s.get('payout_ratio')).number_format = '0.00'
        ws.cell(row=row_idx, column=14, value=s.get('eps')).number_format = '0.00'
        ws.cell(row=row_idx, column=15, value=s.get('dividend_years')).alignment = Alignment(horizontal='center')
        ws.cell(row=row_idx, column=16, value=s.get('roe')).number_format = '0.00'
        ws.cell(row=row_idx, column=17, value=s.get('debt_ratio')).number_format = '0.00'

    # 冻结首行
    ws.freeze_panes = 'A2'

    # 自动列宽
    col_widths = [8, 12, 16, 12, 10, 10, 10, 10, 12, 12, 12, 12, 14, 10, 10, 10, 10]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 输出到内存
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f'红利低波_{data_date}_{count}只.xlsx'
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename,
    )


# ============================================================
# v6.20 配置管理
# ============================================================

@bp.route('/config')
def config_page():
    """配置管理页面"""
    return render_template('config.html')


@bp.route('/api/config', methods=['GET'])
def get_config():
    """
    获取所有配置
    
    返回按分类组织的配置列表
    """
    try:
        config_service = ConfigService.get_instance()
        configs = config_service.get_all()
        
        return jsonify({
            'success': True,
            'data': configs,
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e),
        }), 500


@bp.route('/api/config/batch', methods=['PUT'])
def batch_update_config():
    """
    批量更新配置
    
    Body: {"configs": {"MIN_DIVIDEND_YIELD": "2.5", ...}}
    
    更新后会自动重新运行
    """
    try:
        data = request.get_json()
        configs = data.get('configs', {})
        
        if not configs:
            return jsonify({
                'success': False,
                'error': '未提供配置项',
            }), 400
        
        config_service = ConfigService.get_instance()
        success, message, count = config_service.batch_update(configs)
        
        if not success:
            return jsonify({
                'success': False,
                'error': message,
            }), 400
        
        # 自动重新运行
        try:
            merged = merge_all_data()
            if not merged.empty:
                filtered = filter_stocks(merged)
                if not filtered.empty:
                    scored = calculate_scores(filtered)
                    result = prepare_results(scored)
                    
                    conn = get_db()
                    conn.execute('DELETE FROM stock_data')
                    result.to_sql('stock_data', conn, if_exists='append', index=False)
                    conn.commit()
                    conn.close()
                    
                    return jsonify({
                        'success': True,
                        'message': f'配置已更新，股票池已刷新（{len(result)}只）',
                        'updated_count': count,
                        'pool_size': len(result),
                    })
        except Exception as e:
            print(f"自动运行失败: {e}")
            return jsonify({
                'success': True,
                'message': f'配置已更新（{count}项），但自动运行失败: {str(e)}',
                'updated_count': count,
                'pool_size': 0,
            })
        
        return jsonify({
            'success': True,
            'message': '配置已更新',
            'updated_count': count,
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e),
        }), 500


@bp.route('/api/config/reset', methods=['POST'])
def reset_config():
    """
    恢复默认值
    
    Body: {"category": "A筛选"} 或 {"all": true}
    
    恢复后会自动重新运行
    """
    try:
        data = request.get_json()
        category = data.get('category')
        reset_all = data.get('all', False)
        
        config_service = ConfigService.get_instance()
        
        if reset_all:
            count = config_service.reset_to_default()
            message = f'已恢复所有默认值（{count}项）'
        elif category:
            count = config_service.reset_to_default(category)
            message = f'已恢复{category}的默认值（{count}项）'
        else:
            return jsonify({
                'success': False,
                'error': '请指定 category 或 all',
            }), 400
        
        # 自动重新运行
        try:
            merged = merge_all_data()
            if not merged.empty:
                filtered = filter_stocks(merged)
                if not filtered.empty:
                    scored = calculate_scores(filtered)
                    result = prepare_results(scored)
                    
                    conn = get_db()
                    conn.execute('DELETE FROM stock_data')
                    result.to_sql('stock_data', conn, if_exists='append', index=False)
                    conn.commit()
                    conn.close()
                    
                    return jsonify({
                        'success': True,
                        'message': f'{message}，股票池已刷新（{len(result)}只）',
                        'reset_count': count,
                        'pool_size': len(result),
                    })
        except Exception as e:
            print(f"自动运行失败: {e}")
            return jsonify({
                'success': True,
                'message': f'{message}，但自动运行失败: {str(e)}',
                'reset_count': count,
            })
        
        return jsonify({
            'success': True,
            'message': message,
            'reset_count': count,
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e),
        }), 500


# ==================== v8.0 新增 API ====================

@bp.route('/api/config/strategies', methods=['GET'])
def get_strategies():
    """
    获取预设策略列表
    """
    try:
        config_service = ConfigService.get_instance()
        strategies = config_service.get_preset_strategies()
        current = config_service.get_current_strategy()
        
        return jsonify({
            'success': True,
            'strategies': strategies,
            'current': current,
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e),
        }), 500


@bp.route('/api/config/strategies', methods=['POST'])
def apply_strategy():
    """
    应用预设策略
    
    Body: {"strategy_id": "conservative", "reason": "追求稳定收益"}
    """
    try:
        data = request.get_json()
        strategy_id = data.get('strategy_id')
        reason = data.get('reason', '')
        
        if not strategy_id:
            return jsonify({
                'success': False,
                'error': '请指定 strategy_id',
            }), 400
        
        config_service = ConfigService.get_instance()
        success, message = config_service.apply_preset_strategy(strategy_id, reason)
        
        if not success:
            return jsonify({
                'success': False,
                'error': message,
            }), 400
        
        return jsonify({
            'success': True,
            'message': message,
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e),
        }), 500


@bp.route('/api/config/history', methods=['GET'])
def get_config_history():
    """
    获取参数历史记录
    
    Query: ?key=xxx&limit=50
    """
    try:
        key = request.args.get('key')
        limit = request.args.get('limit', 50, type=int)
        
        config_service = ConfigService.get_instance()
        history = config_service.get_config_history(key, limit)
        
        return jsonify({
            'success': True,
            'history': history,
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e),
        }), 500
